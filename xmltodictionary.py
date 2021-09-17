import os
import xmltodict
import json
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

selected_file = input("Copia il nome del report in formato xml: ")
xmlreport = open(selected_file, "r")
# xmlreport = open("Analyze_Report_20210915_161327.xml", "r")

report = json.dumps(xmltodict.parse(xmlreport.read()), indent=4)
json_report = json.loads(report)

filenames = []
precision = []
recall = []
f_measure = []

for file in json_report["report"]["files"]["file"]:
    filenames.append(os.path.basename(file["path"]))
    for p in file["properties"]["property"]:
        if p["key"] == "EXTRACTION_HITS" and p["value"] == "0":
            precision.append("")
            recall.append("")
            f_measure.append("")
        else:
            if p["key"] == "EXT_PRECISION":
                precision.append(round(float(p["value"]), 2))
            elif p["key"] == "EXT_RECALL":
                recall.append(round(float(p["value"]), 2))
            elif p["key"] == "EXT_F1":
                f_measure.append(round(float(p["value"]), 2))

print(len(filenames))
print(len(precision))
print(len(recall))
print(len(f_measure))

wb = load_workbook("xmlreport.xlsx")
ws = wb.active

for row in range(1, len(filenames)+1):
    for col in range(1, 5):
         char = get_column_letter(col)
         if char == "A":
             ws[char + str(row)] = filenames[row - 1]
         if char == "B":
             ws[char + str(row)] = precision[row - 1]
         if char == "C":
             ws[char + str(row)] = recall[row - 1]
         if char == "D":
             ws[char + str(row)] = f_measure[row - 1]


wb.save("xmlreport.xlsx")